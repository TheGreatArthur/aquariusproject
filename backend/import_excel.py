"""
Script de création et d'initialisation de la base
"""

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import select

from config import DSN, EXCEL_FILE
from models import Poisson, Famille, Genre, ZoneGeo, Robustesse, Comportement, Dispo, Base, TypeEau, ModeVie, Courant
from models.meta import get_engine
from utils import get_or_create_id, get_or_create

engine = get_engine(DSN)

if __name__ == '__main__':
    # Création
    try:
        Base.metadata.create_all(bind=engine)
        exit()
    except:
        pass

    try:
        wb = load_workbook(filename=EXCEL_FILE)
    except PermissionError:
        print("Vous devez fermer Excel d'abord")
        exit()
    except FileNotFoundError:
        print("Classeur Excel non trouvé !")
        exit()

    ws = wb.active

    with Session(engine) as db:
        for row in ws.iter_rows(min_row=2):
            params = dict(
                nom_scientifique=row[1].value,
                nom_commun=row[2].value,
                id_famille=get_or_create_id(db, Famille, nom=row[3].value),
                id_genre=get_or_create_id(db, Genre, nom=row[4].value),
                ph_mini=row[5].value,
                ph_maxi=row[6].value,
                kh=row[7].value,
                gh_mini=row[8].value.split()[0],
                gh_maxi=row[8].value.split()[2],
                taille=float(row[9].value.replace('cm', '')),
                id_zone_geo=get_or_create_id(db, ZoneGeo, nom=row[10].value.strip()),
                nb_individus=row[11].value,
                id_type_eau=get_or_create_id(db, TypeEau, nom=row[12].value.strip()),
                regime=row[13].value,
                id_mode_vie=get_or_create_id(db, ModeVie, nom=row[14].value.strip()),
                temp_mini=int(row[15].value.replace('°C', '')),
                temp_maxi=int(row[16].value.replace('°C', '')),
                id_robustesse=get_or_create_id(db, Robustesse, nom=row[17].value.strip()),
                id_comportement=get_or_create_id(db, Comportement, nom=row[18].value.strip()),
                id_dispo=get_or_create_id(db, Dispo, nom=row[19].value.strip()),
                longevite=int(row[20].value.replace('ans', '')),
                litrage_mini=int(row[21].value.replace('L', '')),
                id_courant=get_or_create_id(db, Courant, nom=row[22].value.strip()) if row[23].value else None,
                points=row[23].value,
            )
            poisson = db.scalar(select(Poisson).filter_by(nom_scientifique=params['nom_scientifique']))
            if poisson:
                print('Edition poisson', poisson)
                for k,v in params.items():
                    if getattr(poisson, k) != v:
                        print(k, ':', getattr(poisson, k), '->', v )
                        setattr(poisson, k, v)
            else:
                db.add(Poisson(**params))
                print('Création poisson', params['nom_scientifique'])

        db.commit()
